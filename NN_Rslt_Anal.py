#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 29 10:03:08 2022

@author: biener
"""

import os
import torch
import numpy as np
import mAECell_Utils as utils
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
#%%
class NN_Stat:
    nAttempts: int = 100
    Data: np.array = None
    Sampled_Data: np.array = None
    Data_Type: str = None
    
    def __init__(self, nAttempts: int = 100, Data: np.array = None, Sampled_Data: np.array = None, Data_Type: str = None):
        self.Ana_O     = NN_Rslt_Anal()
        self.relu1     = utils.ReLU1()
        self.nAttempts = nAttempts
        self.Data      = Data
        self.Sampled_Data = Sampled_Data
        self.Data_Type = Data_Type
        
    def Assemble_Data(self, List_fNames_fName: str = None, Line_fromFile: int = None):
        if List_fNames_fName is None:
            List_fNames_fName = utils.fOpen(Title='Load list of files')
            
        if Line_fromFile is None:
            Line_fromFile = 2999
        
        List_fNames = np.loadtxt(List_fNames_fName, str, delimiter=' ')
        Parameters  = np.loadtxt(List_fNames[0], delimiter=',')
        nDim        = len(Parameters.shape)
        if nDim == 1:
            nParameters = 1
        else:
            nParameters = Parameters.shape[1]
        nFiles      = List_fNames.shape[0]
        Multi_fParmas_Array = np.array(np.zeros((nFiles, nParameters)))
        for ii in range(nFiles - 1):
            if nDim == 1:
                Multi_fParmas_Array[ii] = Parameters[Line_fromFile]
            else:
                Multi_fParmas_Array[ii][:] = Parameters[Line_fromFile][:]
            Parameters  = np.loadtxt(List_fNames[ii + 1], delimiter=',')
        if nDim == 1:
            Multi_fParmas_Array[ii + 1] = Parameters[Line_fromFile]
        else:
            Multi_fParmas_Array[ii + 1][:] = Parameters[Line_fromFile][:]
            
        Multi_fParams_fName = utils.fSaveAs(Title='Save data to:')
        np.savetxt(Multi_fParams_fName, Multi_fParmas_Array, delimiter=',', fmt='%1.7f')
        self.Data = Multi_fParmas_Array
        
        return Multi_fParmas_Array

    def Sample_Data(self, Complete_Data: np.array = None, Data_Points: np.array = None):
        if Complete_Data is None:
            Data_fName    = utils.fOpen(Title='Load complete data')
            Complete_Data = np.loadtxt(Data_fName, delimiter = ',')
        Value = Complete_Data[:,1:]
        Index = Complete_Data[:,0]
        Index = np.exp(Index)

        if Data_Points is None:
            GT_fName    = utils.fOpen(Title='Load ground truth')
            Data_Points = np.loadtxt(GT_fName)[:,0]

        nData_Points  = len(Data_Points)
        Sampled_Value = np.zeros((nData_Points, Value.shape[1]))
        DP            = []
        for ii in range(nData_Points):
            difference_array     = np.absolute(Index - Data_Points[ii])
            indConcTS            = difference_array.argmin()
            DP.append(indConcTS)
            Sampled_Value[ii, :] = Value[indConcTS,:]
            
        self.Sampled_Data = Sampled_Value
        return Sampled_Value, Data_Points
        
    def Weight_Res(self, Ground_Truth: np.array = None, Data_Points: np.array = None, Normalize: bool = True, 
                  reArrange_Data: bool = True):
        if Ground_Truth is None:
            GT_fName     = utils.fOpen(Title='Load ground truth')
            Ground_Truth_plusCoord = np.loadtxt(GT_fName, delimiter = ',')
            Ground_Truth          = Ground_Truth_plusCoord[:,1:]
            if Data_Points is None:
                Data_Points = Ground_Truth_plusCoord[:,0]
            
        if self.Sampled_Data is None:
            self.Sample_Data(Data_Points = Data_Points)

        # Normalize data to put it on the same footage. good for concentration.        
        if Normalize:
            Norm_GT_Factor   = Ground_Truth.sum(axis=1).mean()
            normGT           = Ground_Truth/Norm_GT_Factor
            Norm_Data_Factor = self.Sampled_Data.sum(axis=1).mean()
            normSampled_Data = self.Sampled_Data/Norm_Data_Factor
        else:
            normGT           = Ground_Truth
            normSampled_Data = self.Sampled_Data
        
        # In case concentration is tested and the predicted intermediates were flipped then 
        # the predicted data is rearrange to fit the ground truth.
        if reArrange_Data:
            maxIndex         = np.argmax(normSampled_Data, axis=0)
            Shuffle_iMax     = np.argsort(maxIndex)
            normSampled_Data = normSampled_Data[:,Shuffle_iMax]
        # Find Minimum of predicted data excluding zeros. more effective in the case of concentrations.    
        Data_woZeros     = normSampled_Data + (normSampled_Data==0.0)*1.0
        minSampled_Data  = np.abs(Data_woZeros).min()
        
        WeiRes = (normSampled_Data - normGT)**2/(np.abs(normSampled_Data + (normSampled_Data==0.0)*minSampled_Data))
        WeiRes = WeiRes.sum()
        return WeiRes
    
    def Residual(self, Ground_Truth: np.array = None, Data_Points: np.array = None, Normalize: bool = True, 
                 reArrange_Data: bool = True):
        if Ground_Truth is None:
            GT_fName               = utils.fOpen(Title='Load ground truth')
            Ground_Truth_plusCoord = np.loadtxt(GT_fName, delimiter = ',')
            Ground_Truth           = Ground_Truth_plusCoord[:,1:]
            if Data_Points is None:
                Data_Points = Ground_Truth_plusCoord[:,0]
            
        if self.Sampled_Data is None:
            self.Sample_Data(Data_Points = Data_Points)
        
        # Normalize data to put it on the same footage. good for concentration.
        if Normalize:
            Norm_GT_Factor   = Ground_Truth.sum(axis=1).mean()
            normGT           = Ground_Truth/Norm_GT_Factor
            Norm_Data_Factor = self.Sampled_Data.sum(axis=1).mean()
            normSampled_Data = self.Sampled_Data/Norm_Data_Factor
        else:
            normGT           = Ground_Truth
            normSampled_Data = self.Sampled_Data
        
        # In case concentration is tested and the predicted intermediates were flipped then 
        # the predicted data is rearrange to fit the ground truth.
        if reArrange_Data:
            maxIndex         = np.argmax(normSampled_Data, axis=0)
            Shuffle_iMax     = np.argsort(maxIndex)
            normSampled_Data = normSampled_Data[:,Shuffle_iMax]
            
        Res = (((normSampled_Data-normGT)**2).mean())**0.5
        
        return Res
    
    def Multi_Res(self, List_fNames_fName: str = None, reSample: bool = True, Normalize: bool = True, 
                  reArrange_Data: bool = True, iSelect: int = -1, Weighted_Res: bool = True):
        
        # Load ground truth
        Ground_Truth_fName = utils.fOpen(Title='Load Ground Truth')
        Ground_Truth_plusCoord = np.loadtxt(Ground_Truth_fName, delimiter = ',')
        Ground_Truth       = Ground_Truth_plusCoord[:,1:]
        Data_Points        = Ground_Truth_plusCoord[:,0]
        
        # Load predicted data
        if List_fNames_fName is None:
            List_fNames_fName = utils.fOpen(Title='Load list of files')
        
        List_fNames = np.loadtxt(List_fNames_fName, str, delimiter=' ')
        nFiles      = List_fNames.shape[0]
        Data        = np.loadtxt(List_fNames[0], delimiter = ',')
        
        if reSample:
            self.Sample_Data(Complete_Data = Data, Data_Points = Data_Points)
        else:
            self.Sampled_Data = Data
            
        # calculate Weighted Residual
        if Weighted_Res:
            Merit = self.Weight_Res(Ground_Truth, Data_Points, Normalize, reArrange_Data)
        else:
            Merit = self.Residual(Ground_Truth, Data_Points, Normalize, reArrange_Data)
        
        # Calculate weighted residual for the rest of the files
        for ii in range(nFiles - 1):
            Data = np.loadtxt(List_fNames[ii + 1], delimiter=',')
            if reSample:
                self.Sample_Data(Complete_Data = Data, Data_Points = Data_Points)
            else:
                self.Sampled_Data = Data

            # calculate Weighted Residual
            if Weighted_Res:            
                Merit = np.append(Merit, (self.Weight_Res(Ground_Truth, Data_Points, Normalize, reArrange_Data)))
            else:
                Merit = np.append(Merit, (self.Residual(Ground_Truth, Data_Points, Normalize, reArrange_Data)))
        
        return Merit
    
    def genPred_DED(self, C_File_Path: str = None, Data_Points: torch.Tensor = None, iSelect: int = -1):
        
        if C_File_Path is None:
            C_File_Path = utils.fOpen(Title='Load concentrations file name')
        C_File_Name = os.path.basename(C_File_Path)
        C_File_Dir  = os.path.dirname(C_File_Path)
         
        Scale_File_Path = os.path.join(C_File_Dir, 'InitScale' + C_File_Name[1:]).replace('/','\\')
        Scale_Array = np.loadtxt(Scale_File_Path)
        Scale       = Scale_Array[iSelect]
        c0          = torch.tensor([Scale, 0.0, 0.0, 0.0])
        dTau        = 0.003
        
        if Data_Points is None:
            Data_Points_File_Name  = utils.fOpen(Title='Load time points')
            loaded_array = np.loadtxt(Data_Points_File_Name, usecols=0, dtype='str')
            nTPoints     = int(loaded_array[0])
            Data_Points  = loaded_array[1:nTPoints + 1].astype(float)
        self.Ana_O.tPoints = Data_Points
        
        K_File_Path     = os.path.join(C_File_Dir, 'K' + C_File_Name[1:]).replace('/','\\')
        self.Ana_O.K    = np.exp(np.loadtxt(K_File_Path, delimiter = ','))
        Conc_Interp, tT = self.Ana_O.genPredConc(selEpoch = iSelect, c0 = c0, dTau = dTau, c2logk = False)

        Conc_Interp     = self.relu1(Conc_Interp).double()
        Sampled_C,_     = utils.getConc(Data_Points, tT, Conc_Interp)
        
        nInterm      = Sampled_C.shape[0] - 1
        I_File_Path  = os.path.join(C_File_Dir, 'I' + C_File_Name[1:]).replace('/','\\')
        I            = np.loadtxt(I_File_Path, delimiter=',')
        I_Tensor     = torch.from_numpy(np.reshape(I,(I.shape[0], -1, nInterm)))
        predE        = torch.matmul(I_Tensor[:, iSelect, :].double(), Sampled_C[:-1,:].double())
        self.Sampled_Data = predE
         
        return predE, Data_Points
     
    def Multi_Res_DED(self, List_fNames_fName: str = None, Normalize: bool = False, 
                  reArrange_Data: bool = False, iSelect: int = -1, Weighted_Res: bool = False):
        
        # Load ground truth
        Ground_Truth_fName = utils.fOpen(Title='Load ground truth')
        Ground_Truth_plusCoord = np.loadtxt(Ground_Truth_fName)
        Ground_Truth       = Ground_Truth_plusCoord[:,1:]
        Data_Points        = Ground_Truth_plusCoord[:,0]
        
        # Load predicted data
        if List_fNames_fName is None:
            List_fNames_fName = utils.fOpen(Title='Load list of files')
        
        List_fNames = np.loadtxt(List_fNames_fName, str, delimiter=' ')
        nFiles      = List_fNames.shape[0]
        _, tPoints = self.genPred_DED(List_fNames[0])
            
        # calculate Weighted Residual
        if Weighted_Res:
            Merit = self.Weight_Res(Ground_Truth, Data_Points, Normalize, reArrange_Data)
        else:
            Merit = self.Residual(Ground_Truth, Data_Points, Normalize, reArrange_Data)
        
        # Calculate weighted residual for the rest of the files
        for ii in range(nFiles - 1):
            self.genPred_DED(List_fNames[ii + 1], tPoints)

            # calculate Weighted Residual
            if Weighted_Res:
                Merit = np.append(Merit, (self.Weight_Res(Ground_Truth, Data_Points, Normalize, reArrange_Data)))
            else:
                Merit = np.append(Merit, (self.Residual(Ground_Truth, Data_Points, Normalize, reArrange_Data)))
            
            print('Run No. ' + str(ii + 1))
        
        return Merit
    
    def genHistogram(self, nBins: int = None, Bins_Range: float = None, range_limits: float = None):
        # Setup bins and determine the bin location for each element for the bins
        if range_limits is None:
            R = [np.min(self.Data),np.max(self.Data)]
        else:
            R = range_limits
        N      = self.Data.shape[-1]
        if nBins is None:
            bins = Bins_Range
            nBins = bins.shape[0]
        else:
            bins = np.linspace(R[0],R[1], nBins + 1)
        data2D = self.Data.reshape(-1,N)
        data2D = data2D.transpose()
        idx    = np.searchsorted(bins, data2D,'right') - 1

        # Some elements would be off limits, so get a mask for those
        bad_mask = (idx==-1) | (idx==nBins)

        # We need to use bincount to get bin based counts. To have unique IDs for
        # each row and not get confused by the ones from other rows, we need to 
        # offset each row by a scale (using row length for this).
        scaled_idx = nBins*np.arange(data2D.shape[0])[:,None] + idx

        # Set the bad ones to be last possible index+1 : n_bins*data2D.shape[0]
        limit = nBins*data2D.shape[0]
        scaled_idx[bad_mask] = limit

        # Get the counts and reshape to multi-dim
        counts = np.bincount(scaled_idx.ravel(), minlength = limit + 1)[:-1]
        counts.shape = self.Data.transpose().shape[:-1] + (nBins,)
        return counts, bins
    
    def Plot_Histogram(self, Histogram, Bins, Scale: str=None):
        fig   = plt.figure()
        ax    = fig.add_subplot(1, 1, 1)
        
        line, = ax.plot(Bins, Histogram)
        if Scale is not None:
            if 'xlog' in Scale:
                ax.set_xscale('log')
            if 'ylog' in Scale:
                ax.set_yscale('log')
                
    def Calc_Noise(self, GT: torch.Tensor = None, nSelected_S: int = 3):
        if GT is None:
            GT_File      = utils.fOpen(Title='Load ground truth')
            GT_plusCoord = np.loadtxt(GT_File)
            GT           = GT_plusCoord[:,1:]
            GT           = torch.from_numpy(GT).double()
            
        U, S, V  = torch.svd(GT)
        S[:nSelected_S] = 0
        GT_Noise = torch.mm(torch.mm(U, torch.diag(S)), V.T)
        
        return GT_Noise
    
class XL_Tools:
    wb: Workbook()
    
    def __init__(self):
        pass
        
    def Copy_Template(self, tpl_fName: str=None, xls_fName: str=None):
        if tpl_fName is None:
            tpl_fName = utils.fOpen(Title='Load xls template')
        
        if xls_fName is None:
            xls_fName = utils.fSaveAs(Title='Save data to:')

        os.popen('copy ' + os.path.normpath(tpl_fName) + ' ' + os.path.normpath(xls_fName))
        self.wb = load_workbook(filename = os.path.normpath(xls_fName))
        
        return self.wb
    
    def Find_Sheet(self, Sheet_Name: str=None):
        if Sheet_Name is None:
            Sheet_Name = input('Enter Sheet Name: ')
        
        Sheet_Names = self.wb.sheetnames
        if Sheet_Name in Sheet_Names:
            iSheet = Sheet_Names.index(Sheet_Name)
            self.wb._active_sheet_index = iSheet
        else:
            self.wb.create_sheet(Sheet_Name)
            self.wb._active_sheet_index = len(Sheet_Names)
        
        self.ws = self.wb.active
        
        return self.ws
    
    def Write_Data(self, Data, Cell: str=None):
        if Cell is None:
            Cell = 'A1'
        
        Cell_Coord = coordinate_from_string(Cell)
        istrRow    = Cell_Coord[1]
        istrCol    = column_index_from_string(Cell_Coord[0])
        iendRow    = istrRow + Data.shape[0]
        if Data.ndim > 1:
            iendCol    = istrCol + Data.shape[1]
            
        for ii in range(istrRow,iendRow):
            if Data.ndim == 1:
                self.ws.cell(ii, istrCol, Data[ii-istrRow])
            else:
                for jj in range(istrCol,iendCol):
                    self.ws.cell(ii, jj, Data[ii-istrRow][jj-istrCol])
            print(f'{ii}')
                
    def Save_XLS(self, xls_fName: str=None):
        if xls_fName is None:
            xls_fName = utils.fSaveAs(Title='Save data to:')
        
        self.wb.save(os.path.normpath(xls_fName))
        
class NN_Rslt_Anal:
    Model_Mat: torch.Tensor = torch.tensor([[-1.,-1., 0., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 1., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 1., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 1., 0., 0., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0.,-1.,-1., 0.,-1., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 1., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 0., 1., 0.],\
                                            [ 0., 1., 0., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 1., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0.,-1., 0.,-1.,-1., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 0., 0., 1.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 1., 0., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 1., 0., 0., 0., 0., 0.],\
                                            [ 0., 0., 0., 0., 0., 0., 0., 0.,-1.,-1.]])
    Conc_Interp: np.array = None
    Conc_GT:     np.array = None
    Loaded_Conc: np.array = None
    K:           np.array = None
    K_T:         np.array = None
    tPoints:     np.array = None
    
    def __init__(self, Model_Mat: torch.Tensor = None, K: np.array = None, K_T: np.array = None, 
                 tPoints: np.array = None, Loaded_Conc: np.array = None):
        if Model_Mat is not None:
            self.Model_Mat = Model_Mat
        self.K           = K
        self.K_T         = K_T
        self.tPoints     = tPoints
        self.Loaded_Conc = Loaded_Conc
    
    def transKMatrix(self, File_Name: str = None, Exp: bool=False):
        if self.K is None:
            if File_Name is None:
                File_Name = utils.fOpen(Title='Load Rates Stack')
            K = np.loadtxt(File_Name, delimiter=',')
        else:
            K = self.K
        if Exp:
            K = np.exp(K.transpose())
        else:
            K = K.transpose()
        np.savetxt(File_Name[:-4] + '_T.dat', K, delimiter=',', fmt='%1.7f')
        self.K_T = K
        
        return K
    
    def genConc(self, selIter: int = 3000, c0: torch.Tensor = torch.tensor([1.0, 0.0, 0.0, 0.0]),
                dTau: float = 0.003, c2logk: bool = True):
        
        strFName = utils.fSaveAs(Title='Save Conc. Files')

        Conc, tPoints = self.genGTConc()
        C = Conc.detach().numpy()
        C = np.insert(C, 0, tPoints, axis=0)
        np.savetxt(strFName[:-4]+'_GT.dat', C.T, delimiter=',', fmt='%1.7f')
        self.Conc_GT = C

        Conc_t, tT    = self.genPredConc(selIter, c0, dTau, c2logk)
        C = Conc_t.detach().numpy()
        C = np.insert(C, 0, np.log(tT), axis=0)
        np.savetxt(strFName, C.T, delimiter=',', fmt='%1.7f')
        self.Conc_Interp = C
        
        return self.Conc_Interp, tT, self.Conc_GT, tPoints
    
    def genGTConc(self):
        
        if self.tPoints is None:
            TS_fName     = utils.fOpen(Title='Load Time Stamps')
            loaded_array = np.loadtxt(TS_fName, usecols=0, dtype='str')
            nTPoints     = int(loaded_array[0])
            tPoints      = loaded_array[1:nTPoints + 1].astype(float)
            self.tPoints = tPoints
        else:
            tPoints = self.tPoints
            nTPoints = len(tPoints)

        if self.Loaded_Conc is None:
            CProf_fName  = utils.fOpen(Title='Load GT Concentrations')
            loaded_array = np.loadtxt(CProf_fName)
        else:
            loaded_array = self.Loaded_Conc

        tPoints_Tensor = torch.from_numpy(tPoints).float()
        cTS     = loaded_array[:,0]
        cc      = loaded_array[:,1:]
        cc_Tensor      = torch.from_numpy(cc.T).float()
        Conc, _ = utils.getConc(tPoints_Tensor, cTS, cc_Tensor)
        
        return Conc, tPoints
    
    def genPredConc(self, selEpoch: int = 3000, c0: torch.Tensor = torch.tensor([1.0, 0.0, 0.0, 0.0]), 
                    dTau: float = 0.003, c2logk: bool = True):
        
        if self.tPoints is None:
            TS_fName     = utils.fOpen(Title='Load Time Stamps')
            loaded_array = np.loadtxt(TS_fName, usecols=0, dtype='str')
            nTPoints     = int(loaded_array[0])
            tPoints      = loaded_array[1:nTPoints + 1].astype(float)
        else:
            tPoints = self.tPoints
            nTPoints = len(tPoints)
            
        dTauM   = np.mean(np.log(tPoints[1:]) - np.log(tPoints[:-1]))
        TPoints = np.insert(np.log(tPoints), 0, np.log(tPoints[0])-dTauM)

        if self.K is None:
            File_Name = utils.fOpen(Title='Load Rates Stack')
            k         = np.exp(np.loadtxt(File_Name, delimiter = ','))
        else:
            k = self.K
        k  = torch.from_numpy(k).float()
        tT = np.arange(TPoints[0], TPoints[-1] + dTau/2, dTau)

        if c2logk:
            Conc_t, _ = utils.Analytic_ODE_Solv (torch.exp(k[:,selEpoch]), np.exp(tT), c0, self.Model_Mat)
        else:
            Conc_t, _ = utils.Analytic_ODE_Solv (k[:,selEpoch], np.exp(tT), c0, self.Model_Mat)

        return Conc_t, np.exp(tT)
    
    def genIntrm_File(self, selEpoch: int = 3000, nInterm: int = 3):
        strFName = utils.fOpen(Title='Load Electron Density')
        ED       = np.loadtxt(strFName).astype(int)
        strFName = utils.fOpen(Title='Load Intermediate Stack')
        I        = np.loadtxt(strFName, delimiter=',')
        I        = np.reshape(I,(I.shape[0], -1, nInterm+1))
        
        selI     = I[:,selEpoch-1,:]
        strFName = utils.fSaveAs(Title = 'Save Intermediates Files')
        for ii in range(nInterm):
            currI = selI[:,ii].astype(str)
            currI = np.dstack((ED[:,0].astype(str), currI)).squeeze()
            np.savetxt(strFName[:-4] + str(ii+1) + '.dat', currI, delimiter=' ', fmt='%s')